#Program to read fit files of Jovian maps 
#extract strips of the SST

#Raul Morales-Juberias. Feb 2024

## This program has added functionality to include a cursor that saves the position??
iteration = 5

import glob,os,gc

import numpy as np
import matplotlib.pyplot as plt
import matplotlib.cm as cm

from pylab import *
from scipy import ndimage

import sys
import math
import cv2 ## not sure that this is needed? 

import pandas as pd


import samuel_cass_config



import cv2 ## not sure that this is needed? 


from scipy.ndimage import gaussian_filter1d

import pandas as pd

from load_frames import load_frames
from filter_image import kern_sharpen,lap_sharpen, normalize_images, clahe_sharpen

from drift_vel import find_t_space, track_dark_centroid, shift_image

from scipy.optimize import curve_fit
from scipy.signal import find_peaks

whole = True
if whole:
    images_path = samuel_cass_config.images_path
else: 
    images_path = samuel_cass_config.subset



DEG2RADS = np.deg2rad

class BlittedCursor:
    """
    A cross-hair cursor using blitting for faster redraw.
    """
    def __init__(self, ax):
        self.ax = ax
        self.background = None
        self.horizontal_line = ax.axhline(color='cyan', lw=0.8, ls='--')
        self.vertical_line = ax.axvline(color='cyan', lw=0.8, ls='--')
        # text location in axes coordinates
        self.text = ax.text(0.72, 0.9, '', transform=ax.transAxes,color='cyan')
        self._creating_background = False
        ax.figure.canvas.mpl_connect('draw_event', self.on_draw)

    def on_draw(self, event):
        self.create_new_background()

    def set_cross_hair_visible(self, visible):
        need_redraw = self.horizontal_line.get_visible() != visible
        self.horizontal_line.set_visible(visible)
        self.vertical_line.set_visible(visible)
        self.text.set_visible(visible)
        return need_redraw

    def create_new_background(self):
        if self._creating_background:
            # discard calls triggered from within this function
            return
        self._creating_background = True
        self.set_cross_hair_visible(False)
        self.ax.figure.canvas.draw()
        self.background = self.ax.figure.canvas.copy_from_bbox(self.ax.bbox)
        self.set_cross_hair_visible(True)
        self._creating_background = False

    def on_mouse_move(self, event):
        if self.background is None:
            self.create_new_background()
        if not event.inaxes:
            need_redraw = self.set_cross_hair_visible(False)
            if need_redraw:
                self.ax.figure.canvas.restore_region(self.background)
                self.ax.figure.canvas.blit(self.ax.bbox)
        else:
            self.set_cross_hair_visible(True)
            # update the line positions
            x, y = event.xdata, event.ydata
            self.horizontal_line.set_ydata([y])
            self.vertical_line.set_xdata([x])
            self.text.set_text(f'x={x:1.2f}, y={y:1.2f}')

            self.ax.figure.canvas.restore_region(self.background)
            self.ax.draw_artist(self.horizontal_line)
            self.ax.draw_artist(self.vertical_line)
            self.ax.draw_artist(self.text)
            self.ax.figure.canvas.blit(self.ax.bbox)

def graph2cent(planet,lat):
	import numpy as np

	Jupiter_re     = 71492.e3
	Jupiter_rp     = 66852.e3

	
	latr = lat * DEG2RADS
	obl  = Jupiter_re/Jupiter_rp
	obl2 = obl*obl
	
	latc = 180*np.arctan(np.tan(latr)/obl2)/np.pi
	
	return latc	

def onclick(event):
    global coords, FRAME_IDS, clicked

    if event.xdata is None or event.ydata is None:
        print("No click")
        return
    ix, iy = event.xdata, event.ydata
    print(f'x = {ix}, y = {iy}')

    coords.append([current_frame, ix, iy])  

    clicked = True
    

def find_element(array,element):
        import numpy as np
        idx  = (np.abs(array-element)).argmin()
        return idx

def format_coord(x, y):
    col = find_element(lon_crop, x)  # longitude pixel index
    row = y  # frame / row index
    return f"row={row}, col={col}"


images_list,num_images,images_shape,frames_list = load_frames()


FRAME_IDS = []
coords = []

#for i in range(num_images):
    #FRAME_IDS.append(frames_list[i])


####################################################################################
SHARPEN1 = True
#Kernal sharpening
if SHARPEN1:
    images_list = kern_sharpen(images_list,strength=5)

####################################################################################
SHARPEN2 = False
#Laplacian sharpening 
if SHARPEN2:
    images_list = lap_sharpen(images_list)

NORMALIZED = False
#Normalize the images
if NORMALIZED:
    images_list = normalize_images(images_list) #Without dividing it is just demeaned




lon_init  = 0
lon_final = 360
lat_init  = -80 # I think this may be wrong
lat_final = 80

LON_MIN = 150
LON_MAX = 200
LAT_MIN = -43.5
LAT_MAX = -28.5
OFFSET  = 0.0
LAT_CENTER = (LAT_MAX+LAT_MIN)/2
LAT_CENTER = -30.7 # more accurate than the avg
LAT_CENTER = LAT_CENTER+OFFSET
dlat = 2 # input what lat pixel range for profile to avg over (-34.5 deg)
bandlat = 30 # input what lat pixel half range for each band to be over

print("LAT_CENTER: ",LAT_CENTER)



lon    = np.linspace(lon_init,lon_final,images_shape[1]) #longitudes in pixel space
lat    = np.linspace(lat_init,lat_final,images_shape[0]) #latitudes in pixel space

# dx is number of degrees every pixel, which is 0.1 for these frames
dx     = 360./images_shape[1]
#print(f"dx{dx}")


#Find lon area to focus on. LONMIN and LONMAX are manually found and entered values
j1 = find_element(lon, LON_MIN)
j2 = find_element(lon, LON_MAX)+1
#print(j1, j2)

# make sure lon crop boundaries are correct
if j1 > j2:
    j1, j2 = j2, j1
    print('something weirds going on. check the lon bounds input')


lon_crop = lon[j1:j2] #cropped space that focuses near the ROI
lon_range = j2 - j1 # pixel length of cropped region
print("lon_range: ", lon_range)



# find what pixel row corresponds to the LAT I want to focus on
lat_index = find_element(lat, LAT_CENTER)
pix_shift = dlat*dx
i1 = lat_index - dlat
#print("i1: ",i1)
i2 = lat_index + dlat + 1
#print("i2: ",i2)
g1 = lat_index - bandlat
g2 = lat_index + bandlat + 1
lat_crop = lat[g1:g2]
lat_range = g2-g1
#print(lat_range)


LON_SCANS = np.empty((num_images,lon_range))
print("LON_SCNAS shape: ",np.shape(LON_SCANS))
# initialize contiainer for cropped images which focus on latitude window, width is of cropped region in pixel space
band = [img[i1:i2,j1:j2] for img in images_list]
for i in range(num_images):
    profile = np.mean(band[i],axis=0)
    #print(np.shape(profile))
    profile = (profile-profile.mean())/profile.std()
    LON_SCANS[i,:] = profile

# corresponds frames to time gone by with periods file
# T is periods gone by every frame, tper is hours gone by every frame, DeltaP is periods between frames, deg_omega is ang vel of Jupiter
T,tper,DeltaP, deg_omega = find_t_space(images_list,images_shape)




# track dark vortex of resulting howmoller in both lon and pixel space
xpix_track, xlon_track = track_dark_centroid(
    LON_SCANS,
    lon_crop,
    seed_lon=155,
    dx=dx,
    LON_MIN=150.0,
    LON_MAX=200.0,
    search_halfwidth=10,
    smooth_sigma=1
)


# shift band
shift_from_start = xpix_track - xpix_track[0]
images_list = shift_image(images_list,shift_from_start)
#LON_SCANS = shift_image(LON_SCANS,shift_from_start)

c_max = np.max(LON_SCANS)
c_min = np.min(LON_SCANS)




''' ===================================================================================
    sequence image 
    ===================================================================================
'''
font = {'family' : 'Arial',
        'weight' : 'normal',
        'size'   : 12}

matplotlib.rc('font', **font)

#final lon crop
window_s_lon = 153
window_e_lon = 162
h1 = find_element(lon,window_s_lon)
h2 = find_element(lon,window_e_lon)+1
h_lon_crop = lon[h1:h2]
h_lon_range = h2-h1

window = [img[g1:g2,h1:h2] for img in images_list]
#window1 = kern_sharpen(window,strength=20)
#window2 = lap_sharpen(window)
window1 = clahe_sharpen(window, cL=40.0,n=8)
#window4 = normalize_images(window)
tc = []


print(h_lon_crop[0],h_lon_crop[-1],lat_crop[0],lat_crop[-1])

plt.figure()
plt.imshow(LON_SCANS,cmap='gray',vmin=c_min,vmax=c_max,extent=[lon_crop[0],lon_crop[-1],0,num_images],interpolation='none',origin='lower', aspect='auto')
plt.title("Unshifted. Latitude = "+str(np.round(LAT_CENTER,decimals=1)))
plt.plot(xlon_track, np.arange(len(xlon_track)), 'r-')
#plt.plot(xfit_deg, np.arange(len(xlon_track)), 'b-')
plt.xlabel("Longitude")
plt.ylabel("i")

Re = 71492e3 # Jov equatorial radius in meters
Rp = 66854e3 # Jov polar radius in meters
lam0 = DEG2RADS(0.5*(window_s_lon+window_e_lon))
phi0 = DEG2RADS(LAT_CENTER)
Rj = Re * Rp / np.sqrt((Re * np.sin(phi0))**2 + (Rp * np.cos(phi0))**2) # still in meters


print(lam0,phi0)
print(f"Rj : {Rj:.5e}")

#start = 0

start = len(frames_list)-5
start = 56
#start = 75
# for i in range(3): ##This if for testing things out
for i in range(num_images-5):
    if i >=start:
        print("iteration: ", (i - start))
        current_frame = frames_list[i]
        FRAME_IDS.append(current_frame)
        print(current_frame)

        tc.append(tper[i])

        
        fig = plt.figure(figsize=(12,8), facecolor='white', ) #The units for size here are inches
        fig.canvas.manager.set_window_title(f'Frame {frames_list[i]}')


        plt.subplots_adjust(left=0.08, 
                            bottom=0.01,  
                            right=0.99,  
                            top=0.99,  
                            wspace=0.0,  
                            hspace=0.0)

        ax = fig.add_subplot(1,1,1) 

        c_max = np.max(window[i])
        c_min = np.min(window[i])
        if math.isnan(c_min) or math.isnan(c_max):
            print("Error: One of the values is NaN.")
        else:
            #print("cmax,min",c_max, c_min)
            print("------------")
        plt.imshow(window1[i],cmap='gray',vmin=c_min,vmax=c_max,extent=[h_lon_crop[0],h_lon_crop[-1],lat_crop[0],lat_crop[-1]],interpolation='none',origin='lower',aspect='equal')
        plt.scatter(np.rad2deg(lam0), np.rad2deg(phi0),
            color="red", marker="+", s=120, linewidths=2,
            label="origin")
        date_string = frames_list[i]
        plt.xlim(h_lon_crop[0],h_lon_crop[-1])
        plt.ylim(lat_crop[0],lat_crop[-1])
        plt.title(f"Frame{frames_list[i]}")
        plt.grid("True")
        plt.grid(which='major', color='b', linestyle=':')
        plt.grid(which='minor', color='r', linestyle=':')
        plt.legend
        ax = plt.gca()
        #ax.format_coord = format_coord
        blitted_cursor = BlittedCursor(ax)
        fig.canvas.mpl_connect('motion_notify_event', blitted_cursor.on_mouse_move)
        clicked = False

        cid = fig.canvas.mpl_connect('button_press_event', onclick)

        

        plt.waitforbuttonpress()

        fig.canvas.mpl_disconnect(cid)
        if not clicked:
            coords.append([current_frame, np.nan, np.nan])
            print("NO CLICK")

        plt.show()

        plt.close()
        fig.clf()
        del fig

        gc.collect()

# plt.show()
# plt.savefig('/Users/raul/Desktop/cassini_maps_sequence.png', bbox_inches='tight')
#print(tc)
#print("coords shape: ", np.shape(coords))




# clicks to data analysis
#----------------------------------------------------------
Period=9.92492 # number of hours per period, according to https://ntrs.nasa.gov/citations/19770042742
# extract click data from coords list
coords_arr = np.array(coords, dtype=float)
#print("coords : ",coords_arr)
frame = coords_arr[:,0].astype(int)
lam   = DEG2RADS(coords_arr[:,1]) # in deg
#print(lon)
phi   = DEG2RADS(coords_arr[:,2])

# convert from lon/lat to x/y with set reference frame, use jovian radius at that latitude. centric or graphic?? 
x = Rj * np.cos(phi0)*(lam - lam0) # in m
y = Rj * (phi - phi0) # in m

# convert to polar coords
r = np.sqrt(x**2 + y**2) # in m
theta = np.arctan2(y, x) # in radians
#print("r: ",r)
print("theta: ",theta)

# unwrap so theta does not return to 0 at every cycle
theta_unwrapped = theta.copy()

# create and apply masks so nan values are not present in data
valid = ~np.isnan(theta) 
mask = ~np.isnan(theta)
theta_unwrapped[mask] = np.unwrap(theta[mask])
print("theta_unwrapped: ",theta_unwrapped)
NN = ~mask
theta_v = theta_unwrapped[valid]
tc = np.array(tc, dtype=float)
t_v = tc.copy()
t_v[NN] = np.nan
t_v = t_v[valid] # only non-nan values are considered

print("tc: ", tc)
print("tv: ",t_v)


# fit theta_unwrapped vs time
#----------------------------------------------------------
from scipy.stats import linregress
theta_slope, theta_intercept, theta_r, theta_p, theta_stderr = linregress(t_v, theta_v)
#theta_slope = theta_slope
omega_from_thetafit = -theta_slope / 3600   # rad/s, since t_v is in hours
theta_R2 = theta_r**2

print(f"Omega from theta fit: {omega_from_thetafit:.3e} rad/s")
print(f"Theta fit R^2: {theta_R2:.4f}")

# smooth line for fit
t_fit = np.linspace(t_v[0], t_v[-1], 200)
theta_fit = theta_intercept + theta_slope * t_fit

plt.figure(figsize=(7,5))
plt.plot(t_v, theta_v, 'o', label='Theta data')
plt.plot(t_fit, theta_fit, 'r-', label='Linear fit')

plt.text(
    0.04, 0.20,
    fr'$\omega$ from fit: {omega_from_thetafit:.2e} rad/s' + '\n' +
    fr'$R^2$: {theta_R2:.4f}',
    transform=plt.gca().transAxes,
    va='top',
    ha='left'
)

plt.xlabel('Time (hr)')
plt.ylabel(r'$\theta$ (rad)')
plt.title(f'Theta vs Time. Iteration {iteration}')
plt.legend()
plt.savefig(samuel_cass_config.save_path +f"06-26/Manual/ws_manual_theta_vs_t_{iteration}.pdf", bbox_inches='tight')

#plt.show()





# try fitting position vs time sin curve
#----------------------------------------------------------
from scipy.optimize import curve_fit
from scipy.signal import find_peaks


def sinusoid(x, A, B, C, D):
    return A * np.sin(B * x + C) + D


# using longitude 
lamv = lam[valid] # apply masks to only consider non-nan data points 
mean_lamv = np.mean(lamv)
print("lamv mean: ", mean_lamv)


print("lamv std: ", np.std(lamv))
print(np.max(lamv), np.min(lamv))


A0 = 0.5 * (np.max(lamv) - np.min(lamv))
B0 = 1.6e-5 * 3600  # in rad/hour
C0 = 0
D0 = mean_lamv
initial_guesses = [A0,B0,C0,D0]


t_data = t_v - t_v[0] # take away messy phase shift



popt, pcov = curve_fit(sinusoid,t_data,lamv,p0 = initial_guesses)

A, B, C, D = popt[:] 

print(f"A: {A0:.3f},{A:.3f}")
print(f"B: {B0:.3f},{B:3f}")
print(f"C: {C0:.3f},{C:.3f}")
print(f"D: {D0:.3f},{D:.3f}")

omega_from_lonfit = B / ( 3600) # ang freq in rad / s from fit


# create smooth fit
N = len(t_data)
print(f"N: {N}")
t_graph = np.linspace(t_data[0],t_data[-1], N*20) # create 20 points between each frame
lam_graph = sinusoid(t_graph, *popt) # create fitted lam data 




# find period + ang freq of WS rotation from peaks and troughs
#----------------------------------------------------------
prom = 0.001
peaks, _ = find_peaks(lamv, prominence=prom)
troughs, _ = find_peaks(-lamv,prominence=prom)
lp = lamv[peaks] # lamda of peaks
tp = t_data[peaks] # time of peaks
lt = lamv[troughs] # lambda of troughs
tt  = t_data[troughs] # time of troughs

# find avg period from peaks and troughs. 
if len(peaks) >= 2 and len(troughs) >= 2:

    T_peaks = np.diff(tp)
    T_troughs = np.diff(tt)
    print(np.mean(T_peaks))
    print(np.mean(T_troughs))
    T_mean = 0.5 * (np.mean(T_peaks) + np.mean(T_troughs)) * 3600 # in seconds
    omega_from_peaks = 2*np.pi / T_mean
else:
    omega_from_peaks = np.nan
print("Omega from peaks: ",omega_from_peaks)

plt.figure()

plt.plot(t_v, theta_v, "o-")
plt.plot(tc, theta_unwrapped, "x")
#plt.show()





plt.figure()
plt.plot(t_data,lamv, label="Data")
plt.plot(t_graph,lam_graph, label="Fit", color="g")
plt.scatter(tp, lp, color="red", marker="x", s=50, label="Peaks", zorder=3)
plt.scatter(tt, lt, color="blue", marker="x", s=50, label="Troughs", zorder=3)

plt.text(7, 2.79, fr"Fit: $\omega = {omega_from_lonfit:.2e}$ rad/s", size = 8) 
plt.text(7, 2.78, fr"Peak/Trough distance: $\omega = {omega_from_peaks:.2e}$ rad/s", size = 8) 

plt.title(f"WS Longitude vs Time. Frames 0065-0092. Iteration {iteration}")
plt.xlabel(" Time (hours)")
plt.ylabel(" Longitude (rad)")
plt.ylim(top=2.80)
plt.legend()
plt.savefig(samuel_cass_config.save_path+f"06-26/Manual/ws_manual_lon_vs_t_{iteration}.pdf",bbox_inches="tight")
#plt.show()
#----------------------------------------------------------




# theta step, time step analysis
#----------------------------------------------------------
# take difference between each pair of data points
dtheta = np.diff(theta_v)

# ensure no nan's are still in data. also, I'm not considering any ccw motion.
nan_list = []
dtheta_l = []
for d in dtheta:
    if d >=0 or math.isnan(d):
        dtheta_l.append(math.nan)
        nan_list.append(d)
    elif d<0:
        dtheta_l.append(d)
#for i in range(dtheta):

# diagnostics
print("dth nan's length: ", len(nan_list)) # should be zero
#print("dth_final: ", dtheta_l)
print("len(tc): ", len(tc))
print("len(tv): ",len(t_v),"len(theta): ", len(theta_v))

# take time step between each pair of data points. Each difference corresponds to the correct pair of frames
dt = np.diff(t_v)
print(len(dt), len(dtheta_l))
#print("dt: ", dt)  

omega = (dtheta_l/dt)/3600 # rad/s. dtheta is in rad and dt is in hours
#print("omega: ", omega)
omega = np.nanmean(omega) # take average between all values
print(omega)
#----------------------------------------------------------



#Save the data for the spot position to excel
from openpyxl import load_workbook


# -----------------------
# Summary sheet: one row per iteration
# -----------------------
df_summary = pd.DataFrame([{
    "iteration": iteration,
    "omega_from_lonfit": omega_from_lonfit,
    "omega_from_peaks": omega_from_peaks,
    "omega_from_thetafit": omega_from_thetafit,
    "theta_R2": theta_R2
}])

# -----------------------
# Data sheet: many rows per iteration
# -----------------------
df_data = pd.DataFrame({
    "iteration": iteration,   # repeated automatically
    "frame": FRAME_IDS,
    "lon": [row[1] for row in coords],
    "lat": [row[2] for row in coords],
    "theta": theta,
    "theta_unwrapped": theta_unwrapped,
    "time": tc # hours
})
if True:
    path = samuel_cass_config.save_path + "06-26/manual_spot_data_1.xlsx"

    if not os.path.exists(path):
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df_summary.to_excel(writer, sheet_name="summary", index=False)
            df_data.to_excel(writer, sheet_name="data", index=False)

    else:
        book = load_workbook(path)

        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:

            # append summary
            startrow_summary = book["summary"].max_row
            df_summary.to_excel(
                writer,
                sheet_name="summary",
                index=False,
                header=True,
                startrow=startrow_summary
            )

            # append data
            startrow_data = book["data"].max_row
            df_data.to_excel(
                writer,
                sheet_name="data",
                index=False,
                header=False,
                startrow=startrow_data
            )




    plt.show()



    print('ALL DONE')


    ''''frame': FRAME_IDS,
        'lon': [row[1] for row in coords],
        'lat': [row[2] for row in coords],
        'x' : x,
        'y' : y
    })'''



